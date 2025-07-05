#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys
import os

if not hasattr(sys, 'real_prefix') and not sys.prefix.endswith('venv'):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    venv_path = os.path.join(script_dir, 'venv', 'Scripts', 'python.exe')
    
    if os.path.exists(venv_path):
        print("自动切换到虚拟环境...")
        os.execv(venv_path, [venv_path] + sys.argv)
    else:
        print("未找到虚拟环境，将在全局环境中运行。")

import argparse
import json
import os
import time
import pandas as pd
import requests
from openpyxl import load_workbook
from tqdm import tqdm

def translate_batch(texts, source_lang, target_lang, config):
    """调用API批量翻译文本，支持重试机制"""
    url = config["api_url"]
    timeout = config["api_timeout"]
    max_retries = config.get("max_retries", 3)  # 默认重试3次
    retry_delay = config.get("retry_delay", 1)  # 默认重试延迟1秒
    
    prompt = config["translation_prompt"]["instruction"].replace(
        "[待翻译的语言]", source_lang).replace(
        "[要翻译成的语言]", target_lang)
    
    requirements = "\n".join(config["translation_prompt"]["requirements"])
    input_json = json.dumps(texts, ensure_ascii=False, indent=2)
    full_content = f"{prompt}\n\n要求:\n{requirements}\n\n输入:\n{input_json}"
    
    payload = {
        "model": config["model_name"],
        "messages": [{"role": "user", "content": full_content}],
        "stream": False,
        "max_tokens": config["max_tokens"],
        "temperature": config["temperature"],
        "top_p": config["top_p"],
        "top_k": config["top_k"],
        "frequency_penalty": config["frequency_penalty"],
        "enable_thinking": config["enable_thinking"]
    }
    
    headers = {
        "Authorization": f"Bearer {config['api_key']}",
        "Content-Type": "application/json"
    }
    
    for attempt in range(max_retries + 1):  # 尝试次数 = 重试次数 + 1
        try:
            response = requests.post(url, json=payload, headers=headers, timeout=timeout)
            response.raise_for_status()
            result = response.json()
            
            # 检查API返回结果
            if 'choices' in result and len(result['choices']) > 0:
                content = result['choices'][0]['message']['content'].strip()
                try:
                    translated_texts = json.loads(content)
                    if isinstance(translated_texts, list) and len(translated_texts) == len(texts):
                        return translated_texts  # 成功返回翻译结果
                    else:
                        # 格式错误但内容可能可用，尝试提取有效部分
                        if isinstance(translated_texts, list) and len(translated_texts) > 0:
                            # 尝试使用部分结果
                            if len(translated_texts) < len(texts):
                                # 补充缺失的部分
                                translated_texts += [f"[部分翻译] {text}" for text in texts[len(translated_texts):]]
                            return translated_texts
                        raise ValueError("返回结果格式不正确")
                except json.JSONDecodeError:
                    # 尝试从非JSON响应中提取翻译内容
                    if "[" in content and "]" in content:
                        try:
                            # 尝试提取可能的JSON数组
                            start_idx = content.index("[")
                            end_idx = content.rindex("]") + 1
                            json_content = content[start_idx:end_idx]
                            translated_texts = json.loads(json_content)
                            if isinstance(translated_texts, list) and len(translated_texts) == len(texts):
                                return translated_texts
                        except:
                            pass
                    raise ValueError(f"API返回内容无法解析为JSON: {content[:200]}...")
            else:
                raise ValueError(f"API返回结果缺少choices字段: {result}")
        
        except (requests.exceptions.RequestException, ValueError, json.JSONDecodeError) as e:
            if attempt < max_retries:
                error_type = type(e).__name__
                print(f"尝试 {attempt+1}/{max_retries+1} 失败 ({error_type}): {str(e)[:200]}")
                print(f"等待 {retry_delay} 秒后重试...")
                time.sleep(retry_delay)
            else:
                print(f"所有重试失败: {str(e)[:200]}")
                if hasattr(e, 'response') and e.response is not None:
                    print(f"错误响应内容: {e.response.text[:500]}")
                return [f"[翻译错误] 重试失败: {str(e)[:100]}"] * len(texts)
        
        except Exception as e:
            print(f"未处理的异常: {str(e)}")
            return [f"[翻译错误] 未处理异常: {str(e)[:100]}"] * len(texts)
    
    return [f"[翻译错误] 未知错误"] * len(texts)

def process_file(file_path, col_index, source_lang, target_lang, config_path, batch_size):
    """处理文件并添加翻译列"""
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    ext = os.path.splitext(file_path)[1].lower()
    output_path = file_path.replace(ext, f"_translate{ext}")
    
    if ext in ['.xlsx', '.xls']:
        wb = load_workbook(file_path)
        ws = wb.active
        
        translation_col = col_index + 1
        ws.insert_cols(translation_col)
        ws.cell(row=1, column=translation_col, value=f"翻译({source_lang}→{target_lang})")
        
        texts_to_translate = []
        cell_positions = []
        
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col_index).value
            if cell_value and str(cell_value).strip():
                texts_to_translate.append(str(cell_value))
                cell_positions.append(row)
        
        for i in tqdm(range(0, len(texts_to_translate), batch_size), desc="批量翻译进度"):
            batch = texts_to_translate[i:i+batch_size]
            translated_batch = translate_batch(batch, source_lang, target_lang, config)
            
            for j, row in enumerate(cell_positions[i:i+batch_size]):
                ws.cell(row=row, column=translation_col, value=translated_batch[j])
            
            time.sleep(config["api_delay"])
        
        wb.save(output_path)
        print(f"Excel文件已保存: {output_path}")
    
    elif ext == '.csv':
        df = pd.read_csv(file_path)
        translation_col = df.columns[col_index-1] + f"_翻译({source_lang}→{target_lang})"
        df.insert(col_index, translation_col, "")
        
        texts_to_translate = []
        row_indices = []
        
        for idx in range(len(df)):
            text = str(df.iloc[idx, col_index-1])
            if text.strip():
                texts_to_translate.append(text)
                row_indices.append(idx)
        
        for i in tqdm(range(0, len(texts_to_translate), batch_size), desc="批量翻译进度"):
            batch = texts_to_translate[i:i+batch_size]
            translated_batch = translate_batch(batch, source_lang, target_lang, config)
            
            for j, idx in enumerate(row_indices[i:i+batch_size]):
                df.iloc[idx, col_index] = translated_batch[j]
            
            time.sleep(config["api_delay"])
        
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        print(f"CSV文件已保存: {output_path}")
    
    else:
        raise ValueError("不支持的文件格式，请使用.xlsx, .xls或.csv文件")

def main():
    parser = argparse.ArgumentParser(description="AI文件翻译工具")
    parser.add_argument("-F", "--file", dest="file", help="输入文件路径 (.xlsx/.xls/.csv)")
    parser.add_argument("-C", "--col", type=int, dest="col", help="待翻译列号 (从1开始)")
    parser.add_argument("-S", "--source", dest="source", help="源语言代码")
    parser.add_argument("-T", "--target", dest="target", help="目标语言代码")
    parser.add_argument("-G", "--config", default="AI_config.json", dest="config", help="AI配置JSON文件路径")
    parser.add_argument("-B", "--batch", type=int, dest="batch", help="一次翻译行数")
    parser.add_argument("-R", "--retries", type=int, dest="retries", 
                        help="API失败重试次数 (默认从配置读取)")
    
    args = parser.parse_args()

    with open(args.config, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    file_path = args.file if args.file else config["default_file"]
    col_index = args.col if args.col else config["default_column"]
    source_lang = args.source if args.source else config["default_source_lang"]
    target_lang = args.target if args.target else config["default_target_lang"]
    batch_size = args.batch if args.batch else config["default_batch_size"]
    
    # 处理重试次数参数
    if args.retries is not None:
        config["max_retries"] = args.retries
    elif "max_retries" not in config:
        config["max_retries"] = 3  # 默认值
    
    if col_index < 1:
        raise ValueError("列号必须大于等于1")
    
    if batch_size < 1 or batch_size > config["max_batch_size"]:
        raise ValueError(f"批量大小必须在1-{config['max_batch_size']}之间")
    
    print(f"开始翻译: {file_path} 第{col_index}列 ({source_lang}→{target_lang})")
    print(f"配置: 批量大小={batch_size}, 最大重试次数={config['max_retries']}")
    process_file(file_path, col_index, source_lang, target_lang, args.config, batch_size)

if __name__ == "__main__":
    print("成功切换虚拟环境！")
    main()
